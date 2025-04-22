from flask import Flask, render_template, request, redirect, flash
import openpyxl
import os

app = Flask(__name__)
app.secret_key = 'your-secret-key'

EXCEL_FILE = 'data.xlsx'

def get_existing_data():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:  # name column
            data.append({'name': row[0], 'email': row[1]})
    return data

@app.route('/', methods=['GET', 'POST'])
def form():
    data = get_existing_data()

    if request.method == 'POST':
        name = request.form['name'].strip()
        email = request.form['email'].strip()

        # Check if name already exists (case-insensitive)
        if any(d['name'].lower() == name.lower() for d in data):
            flash("This name already exists. Please enter a unique name.")
            return redirect('/')

        # Write to Excel
        wb = openpyxl.load_workbook(EXCEL_FILE)
        sheet = wb.active
        next_row = sheet.max_row + 1
        sheet.cell(row=next_row, column=1, value=name)
        sheet.cell(row=next_row, column=2, value=email)
        wb.save(EXCEL_FILE)

        flash("Data saved successfully!")
        return redirect('/')

    return render_template('form.html', names=[d['name'] for d in data], emails=[d['email'] for d in data if d['email']])

@app.route('/submit', methods=['POST'])
def submit():
    selected = request.form.get('person')       # value from dropdown
    new_name = request.form.get('new_name')     # new name input
    new_email = request.form.get('new_email')   # new email input

    if selected:
        name, email = selected.split('|')
    elif new_name and new_email:
        name = new_name
        email = new_email
    else:
        return "Please select a person or enter a new name and email."

    # Save to Excel
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([name, email])
    wb.save(EXCEL_FILE)

    return f"Saved: {name} ({email})"

if __name__ == '__main__':
    app.run(debug=True)